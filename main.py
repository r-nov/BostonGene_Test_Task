import requests
from bs4 import BeautifulSoup
import json
from openpyxl import Workbook, load_workbook
import os.path


TestUniProtIDs = {'sCD40L': 'P29965', 'EGF': 'P01133', 'Eotaxin': 'P51671', 'FGF-2': 'P09038', 'FLT-3L': 'P49771',
              'Fractalkine': 'P78423', 'G-CSF': 'Q9GJU0', 'GM-CSF': 'P04141', 'GRO-a': 'B5X8A6', 'IFN-a2': 'P01563',
              'IFN-g': 'P01579', 'IL-1a':'P01583'} #Dict with keys - proteins names from Test Task and values - UniProtIDs of these proteins

def main(UniProtID: str, format: str):
    '''Return protein info in json or excel format.

    Keyword arguments:
    UniProtID -- protein ID in UniProt DB
    format -- format in which you want to get information about protein (json or excel)

    '''

    url = f'http://www.uniprot.org/uniprot/{UniProtID}.xml'
    ProteinInfo = XMLparser(url)
    if format == 'json':
        JsonProteinInfo = SaveToJSON(ProteinInfo)
        return JsonProteinInfo
    elif format  == 'excel':
        SaveToXLSX(ProteinInfo, UniProtID)

def XMLparser(url: str):
    '''Return information about protein in form of a python dictionary via parsing page about protein in xml format.

    Keyword argument:
    url -- URL of page about protein in UniProt DB in xml format

    '''

    source = requests.get(url).text
    soup = BeautifulSoup(source, 'lxml')

    UniProtID = soup.find('accession').text

    RecFullProteinName = []
    try:
        for fullname in soup.find('recommendedname').find_all('fullname'):
            RecFullName = fullname.text
            RecFullProteinName.append(RecFullName)
    except Exception as e:
        pass

    RecShortProteinName = []
    try:
        for shortname in soup.find('recommendedname').find_all('shortname'):
            RecShortName = shortname.text
            RecShortProteinName.append(RecShortName)
    except Exception as e:
        pass

    SubFullProteinName = []
    try:
        for fullname in soup.find_all('submittedname'):
            SubFullName = fullname.fullname.text
            SubFullProteinName.append(SubFullName)
    except Exception as e:
        pass

    SubShortProteinName = []
    try:
        for shortname in soup.find_all('submittedname'):
            if shortname.shortname:
                SubShortName = shortname.shortname.text
                SubShortProteinName.append(SubShortName)
            else:
                pass
    except Exception as e:
        pass

    AltFullProteinName = []
    try:
        for alternativeName in soup.find_all('alternativename'):
            AltFullName = alternativeName.fullname.text
            AltFullProteinName.append(AltFullName)
    except Exception as e:
        pass

    AltShortProteinName = []
    try:
        for alternativeName in soup.find_all('alternativename'):
            if alternativeName.shortname:
                AltShortName = alternativeName.shortname.text
                AltShortProteinName.append(AltShortName)
            else:
                pass
    except Exception as e:
        pass

    PrimaryGeneName = []
    try:
        for primaryName in soup.gene.find_all('name', type='primary'):
            PrimaryName = primaryName.text
            PrimaryGeneName.append(PrimaryName)
    except Exception as e:
        pass

    SynonymGeneName = []
    try:
        for synonymName in soup.gene.find_all('name', type='synonym'):
            SynonymName = synonymName.text
            SynonymGeneName.append(SynonymName)
    except Exception as e:
        pass

    functions = []
    try:
        for function in soup.find_all('comment', type='function'):
            Function = function.text
            functions.append(Function)
    except Exception as e:
        pass
    Functions = ''.join(functions).replace('\n', ' ').replace('  ', ' ')[1:-1] if functions else None

    ProteinInfo = {
        'UniProtID': UniProtID, 'RecommendedFullProteinName': RecFullProteinName,
        'RecommendedShortProteinName': RecShortProteinName, 'SubmittedFullProteinName': SubFullProteinName,
        'SubmittedShortProteinName': SubShortProteinName, 'AlternativeFullProteinName': AltFullProteinName,
        'AlternativeShortProteinName': AltShortProteinName, 'PrimaryGeneName': PrimaryGeneName,
        'SynonymGeneName': SynonymGeneName, 'Functions': Functions
    }

    for key, value in ProteinInfo.items():
        if value == []:
            ProteinInfo[key] = None
        else:
            pass

    return ProteinInfo

def SaveToJSON(ProteinInfo: dict):
    '''Save dictionary with information about protein in the json object.

    Keyword argument:
    ProteinInfo -- python dictionary with information about protein

    '''

    json_object = json.dumps(ProteinInfo, indent=2)
    return json_object

def SaveToXLSX(ProteinInfo: dict, UniProtID: str):
    '''Save dictionary with information about protein in the excel table (.xlsx extension).

        Keyword argument:
        ProteinInfo -- python dictionary with information about protein

        '''

    if os.path.exists(f'{os.getcwd()}/results.xlsx'):
        pass
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'ProteinsInfo'
        headings = list(ProteinInfo.keys())
        ws.append(headings)
        wb.save('results.xlsx')
    wb = load_workbook('results.xlsx')
    ws = wb.active
    FirstColumn = ws['A']
    UniProtIDs = [FirstColumn[ID].value for ID in range(1, len(FirstColumn))]
    if UniProtID in UniProtIDs:
        pass
    else:
        values = []
        for value in ProteinInfo.values():
            if type(value) == list:
                values.append(', '.join(value))
            else:
                values.append(value)
        ws.append(values)
    wb.save('results.xlsx')

if __name__ == '__main__':
    for ID in TestUniProtIDs.values():
        main(ID, 'excel')


